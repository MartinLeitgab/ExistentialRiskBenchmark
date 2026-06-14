"""Build a verbatim model-response matrix spreadsheet from a prompt_validation run.

One worksheet per scenario. One row per (prompt variation x run). Columns:
  [prompt config] | Claude: choice + verbatim response + 6 Judge-A scores
                  | Gemini: choice + verbatim response + 6 Judge-A scores
                  | GPT:    choice + verbatim response + 6 Judge-A scores

Judge B (reasoning judge) is not yet built (issue #42), so "judge outputs" here
are the Judge-A 6 dimensions only. This is stated explicitly in the header row.

Fail-fast: crashes if raw/judge row counts diverge or a (scenario,variant,model,
run) key is missing its judge join, rather than silently emitting blank cells.
"""

import ast
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Allow `from pipeline_a_scenarios...` import when run from repo root.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from pipeline_a_scenarios.utils.prompt_generator import generate_prompt  # noqa: E402

RUN_DIR = Path(
    "outputs/data_Riccardo060926/results_phase1-rerun_060926/results/prompt_validation"
)
SEEDS_PATH = Path("data/scenarios/seeds_phase1.json")
OUT_PATH = Path("outputs/data_Riccardo060926/model_responses_matrix.xlsx")

# Placeholders substituted for the scenario body in the reconstructed prompt
# column (the full scenario text lives on the Scenarios tab, per request).
_CTX_PH = "[SCENARIO CONTEXT - see Scenarios tab]"
_A_PH = "[ACTION A - see Scenarios tab]"
_B_PH = "[ACTION B - see Scenarios tab]"


def _reconstruct_prompt(dims):
    """Rebuild the exact system+user prompt sent, with scenario body excluded."""
    p = generate_prompt(_CTX_PH, _A_PH, _B_PH, dimensions=dims)
    return f"SYSTEM:\n{p['system_prompt']}\n\nUSER:\n{p['user_prompt']}"


# Display order for the three model columns.
MODELS = ["claude-opus-4-7", "gemini-3.1-pro-preview", "gpt-5.5"]
MODEL_LABEL = {
    "claude-opus-4-7": "Claude",
    "gemini-3.1-pro-preview": "Gemini",
    "gpt-5.5": "GPT-5.5",
}

JUDGE_DIMS = [
    ("preference_signal_score", "PrefSignal"),
    ("strategic_responding_score", "Strategic"),
    ("coherence_score", "Coherence"),
    ("reasoning_depth_score", "ReasonDepth"),
    ("format_compliance_score", "FormatCompl"),
    ("instrumental_vs_terminal_score", "IVT"),
]

DIM_COLS = [
    "framing",
    "exfiltration",
    "response_format",
    "system_prompt",
    "temperature",
    "action_order",
    "ethical_framing",
    "value_conflict",
    "goal_specification",
]


def _parse_dims(raw_val):
    if isinstance(raw_val, dict):
        return raw_val
    try:
        return ast.literal_eval(raw_val)
    except (ValueError, SyntaxError):
        return {}


def main():
    raw = json.load(open(RUN_DIR / "raw_responses.json", encoding="utf-8"))
    judge = json.load(open(RUN_DIR / "judge_metrics.json", encoding="utf-8"))

    if len(raw) != len(judge):
        sys.exit(f"FAIL: raw n={len(raw)} != judge n={len(judge)}")

    def key(r):
        return (r["scenario_id"], r["variant_id"], r["model"], str(r["run"]))

    judge_by_key = {key(j): j for j in judge}
    if len(judge_by_key) != len(judge):
        sys.exit("FAIL: judge rows are not unique on (scenario,variant,model,run)")

    # raw_by_key[(scenario,variant,run)][model] -> raw row
    raw_index = {}
    for r in raw:
        k = key(r)
        if k not in judge_by_key:
            sys.exit(f"FAIL: raw row {k} has no judge join")
        cell = raw_index.setdefault(
            (r["scenario_id"], r["variant_id"], str(r["run"])), {}
        )
        cell[r["model"]] = r

    scenarios = sorted({r["scenario_id"] for r in raw})

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    model_fills = {
        "claude-opus-4-7": PatternFill("solid", fgColor="E2EFDA"),
        "gemini-3.1-pro-preview": PatternFill("solid", fgColor="FCE4D6"),
        "gpt-5.5": PatternFill("solid", fgColor="FFF2CC"),
    }
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    wb.remove(wb.active)

    for sc in scenarios:
        ws = wb.create_sheet(title=sc.replace("proto_", "p")[:31])

        # Build header.
        cfg_headers = ["variant_id", "run"] + DIM_COLS + ["prompt_sent"]
        headers = list(cfg_headers)
        model_block_starts = {}
        for m in MODELS:
            model_block_starts[m] = len(headers) + 1
            lbl = MODEL_LABEL[m]
            headers.append(f"{lbl}_choice")
            headers.append(f"{lbl}_response")
            headers.extend(f"{lbl}_{short}" for _, short in JUDGE_DIMS)
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = bold
            c.fill = header_fill
            c.alignment = wrap
        for m in MODELS:
            ws.cell(row=1, column=model_block_starts[m]).fill = model_fills[m]

        # Rows: one per (variant, run).
        cells = [(v, run) for (s, v, run) in raw_index if s == sc]
        cells = sorted(set(cells))

        for variant_id, run in cells:
            block = raw_index[(sc, variant_id, run)]
            any_raw = next(iter(block.values()))
            dims = _parse_dims(any_raw.get("dimensions"))
            prompt_sent = _reconstruct_prompt(dims) if dims else ""
            row = (
                [variant_id, run] + [dims.get(d, "") for d in DIM_COLS] + [prompt_sent]
            )
            for m in MODELS:
                rr = block.get(m)
                if rr is None:
                    row.extend([""] * (2 + len(JUDGE_DIMS)))
                    continue
                jr = judge_by_key[(sc, variant_id, m, str(run))]
                row.append(rr.get("parsed_choice", ""))
                row.append(rr.get("response_text", ""))
                row.extend(jr.get(field, "") for field, _ in JUDGE_DIMS)
            ws.append(row)

        # Column widths + wrapping for response columns.
        for col_idx, h in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            if h.endswith("_response") or h == "prompt_sent":
                ws.column_dimensions[letter].width = 70
            elif h in ("variant_id",):
                ws.column_dimensions[letter].width = 28
            elif h in DIM_COLS:
                ws.column_dimensions[letter].width = 13
            else:
                ws.column_dimensions[letter].width = 11
        for r in range(2, ws.max_row + 1):
            for col_idx, h in enumerate(headers, start=1):
                if h.endswith("_response") or h == "prompt_sent":
                    ws.cell(row=r, column=col_idx).alignment = wrap
        ws.freeze_panes = "C2"

    # Full-scenario tab.
    seeds = json.load(open(SEEDS_PATH, encoding="utf-8"))
    sc_ws = wb.create_sheet(title="Scenarios")
    sc_fields = [
        "id",
        "preference_pair",
        "pair_type",
        "difficulty",
        "stakes_level",
        "domain",
        "context",
        "current_use_context",
        "action_a",
        "action_a_preference",
        "action_b",
        "action_b_preference",
        "design_notes",
    ]
    sc_ws.append(sc_fields)
    for col in range(1, len(sc_fields) + 1):
        c = sc_ws.cell(row=1, column=col)
        c.font = bold
        c.fill = header_fill
    for s in seeds:
        sc_ws.append([str(s.get(f, "")) for f in sc_fields])
    for col_idx, f in enumerate(sc_fields, start=1):
        letter = get_column_letter(col_idx)
        if f in (
            "context",
            "current_use_context",
            "action_a",
            "action_b",
            "design_notes",
        ):
            sc_ws.column_dimensions[letter].width = 80
            for r in range(2, sc_ws.max_row + 1):
                sc_ws.cell(row=r, column=col_idx).alignment = wrap
        else:
            sc_ws.column_dimensions[letter].width = 18
    sc_ws.freeze_panes = "A2"

    # Read-me sheet.
    info = wb.create_sheet(title="_README", index=0)
    notes = [
        ["Model-response matrix — PIPE-A7 Phase 1 re-run (data_Riccardo060926)"],
        [""],
        ["One worksheet per scenario. One row per (prompt variation x run)."],
        ["Column blocks per model: choice | verbatim response | 6 Judge-A scores."],
        [""],
        [
            "'prompt_sent' column = the exact SYSTEM + USER prompt sent to the model, "
            "reconstructed via prompt_generator.generate_prompt() from the row's "
            "dimensions. The scenario body is replaced by placeholders "
            "([SCENARIO CONTEXT...]/[ACTION A...]/[ACTION B...]); full scenario text "
            "is on the 'Scenarios' tab. Calibration variants show their injected "
            "goal directive inside this column."
        ],
        [
            "'Scenarios' tab: every seed in full (context, current_use_context, both "
            "actions + preference labels, design_notes)."
        ],
        [""],
        [
            "Judge-A dimensions: PrefSignal, Strategic, Coherence, ReasonDepth, "
            "FormatCompl, IVT (1-5)."
        ],
        ["Judge B (reasoning judge, issue #42) is NOT built yet — not included."],
        [""],
        [
            "deployment_context in the source data is mislabeled 'current_use' on "
            "every row (prompt_validation.py:120 short-circuits whenever a scenario "
            "has a current_use_context field). Actual framing is set by the "
            "system_prompt dimension column (auto/asst/ag/absent)."
        ],
        [""],
        [f"Source: {RUN_DIR.as_posix()}"],
        [
            "Rows include calibration variants (icd/ahd/ahdc/phd suffixes) — "
            "filter those out for variant ranking."
        ],
    ]
    for line in notes:
        info.append(line)
    info.column_dimensions["A"].width = 110
    info["A1"].font = Font(bold=True, size=13)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"OK wrote {OUT_PATH}  ({len(scenarios)} scenario tabs)")
    for sc in scenarios:
        n = len({(v, run) for (s, v, run) in raw_index if s == sc})
        print(f"  {sc}: {n} prompt-variation rows")


if __name__ == "__main__":
    main()
